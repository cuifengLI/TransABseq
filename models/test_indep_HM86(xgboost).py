import numpy as np
import os
from sklearn.metrics import matthews_corrcoef, f1_score, roc_curve
from sklearn import metrics
from sklearn.metrics import confusion_matrix, auc, precision_recall_curve
from sklearn.model_selection import ShuffleSplit, StratifiedKFold, train_test_split
import openpyxl as op
import matplotlib.pyplot as plt
from model import get_model
import pandas as pd
from openpyxl import load_workbook
import warnings
from sklearn.metrics import mean_absolute_error, mean_squared_error, r2_score
from scipy.stats import pearsonr, spearmanr
import seaborn as sb
import joblib
import tensorflow as tf

warnings.filterwarnings("ignore")

filename = 'XXX.xlsx'


def op_toexcel(data, filename):
    if os.path.exists(filename):
        wb = op.load_workbook(filename)
        ws = wb.worksheets[0]
        ws.append(data)
        wb.save(filename)
    else:
        wb = op.Workbook()  
        ws = wb['Sheet']  
        ws.append(['MSE', 'MAE', 'RMSE', 'R2', 'PCC', 'P_value'])
        ws.append(data)  
        wb.save(filename)


def evaluate_regression(i, j):
    all_esm = np.lib.format.open_memmap('../features_npy/HM86_6esm.npy')
    all_prot = np.lib.format.open_memmap('../features_npy/HM86_wild_161_prot.npy')
    all_label = np.lib.format.open_memmap('../features_npy/HM86_label.npy')

    all_label = all_label.astype(np.float)
    print(all_label.dtype)

    encoder_model = get_model()
    encoder_model.load_weights(f'XXX.h5')

    feature_layer = tf.keras.models.Model(inputs=encoder_model.input,
                                          outputs=encoder_model.get_layer('concatenate').output)
    test_feature_output = feature_layer.predict([all_esm, all_prot])
    ml_model = joblib.load(f'XXX.pkl')

    y_pred = ml_model.predict(test_feature_output).reshape(-1, )
    y_true = all_label

    mse = mean_squared_error(y_true, y_pred)
    mae = mean_absolute_error(y_true, y_pred)
    rmse = np.sqrt(mse)
    r2 = r2_score(y_true, y_pred)
    pearsonr_corr, p_value = pearsonr(y_true, y_pred)

    print(f"INDE Mean Squared Error (MSE): {mse:.4f}")
    print(f"INDE Mean Absolute Error (MAE): {mae:.4f}")
    print(f"INDE RMSE: {rmse:.4f}")
    print(f"INDE R-squared (R2) Score: {r2:.4f}")
    print(f"INDE PCC: {pearsonr_corr:.4f}")
    print(f"INDE P-value: {p_value:.4f}")

    result = mse, mae, rmse, r2, pearsonr_corr, p_value
    op_toexcel(result, filename)

    df = pd.DataFrame({
        'True': y_true,
        'Predict': y_pred,
    })
 
    filename2 = f'XXX.xlsx'
    try:
        book = load_workbook(filename2)  
        book.remove(book.active)  
    except FileNotFoundError:
        pass  
    df.to_excel(filename2, index=False)


if __name__ == "__main__":
    os.environ["CUDA_VISIBLE_DEVICES"] = "0"
    gpus = tf.config.experimental.list_physical_devices('GPU')
    tf.config.experimental.set_memory_growth(gpus[0], True)
    for i in range(1, 2):
        for j in range(1, 2):
            evaluate_regression(i, j)
